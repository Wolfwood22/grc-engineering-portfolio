#!/usr/bin/env python3
"""
poam_generator.py — POA&M Excel Workbook Generator

Takes STIG parser CSV output and produces a formatted Plan of Action
and Milestones (POA&M) workbook ready for ISSO/AO review or eMASS import.
"""

import sys
from datetime import datetime, timedelta
from pathlib import Path

import click
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.panel import Panel
from rich.table import Table

console = Console()

# ── Remediation timelines by severity (DoD/FedRAMP standard) ─────────────────
REMEDIATION_DAYS = {
    "high":   30,
    "medium": 90,
    "low":    180,
}

# ── Cell fill colors ──────────────────────────────────────────────────────────
FILL_COLORS = {
    "high":   PatternFill("solid", fgColor="FFD7D7"),  # light red
    "medium": PatternFill("solid", fgColor="FFF9C4"),  # light yellow
    "low":    PatternFill("solid", fgColor="D7F5D7"),  # light green
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy
COVER_FILL  = PatternFill("solid", fgColor="2E5F9E")   # medium blue
WHITE_FONT  = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT   = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ── POA&M column definitions: (header label, column width) ───────────────────
POAM_COLUMNS = [
    ("Weakness ID",                    12),
    ("Weakness Name",                  50),
    ("Point of Contact",               22),
    ("Resources Required",             25),
    ("Scheduled Completion Date",      25),
    ("Milestone with Completion Date", 35),
    ("Changes to Milestones",          30),
    ("Source Identifying Weakness",    25),
    ("Status",                         12),
    ("Category",                       10),
    ("Severity",                       12),
    ("Rule ID",                        30),
]


# ── Data layer ────────────────────────────────────────────────────────────────

def load_findings(filepath: str) -> pd.DataFrame:
    """Read STIG parser CSV into a DataFrame and validate required columns."""

    path = Path(filepath)

    if not path.exists():
        console.print(f"[red]Error:[/red] File not found: {filepath}")
        sys.exit(1)

    if path.suffix.lower() != ".csv":
        console.print(f"[red]Error:[/red] Expected a .csv file, got: {path.suffix}")
        sys.exit(1)

    df = pd.read_csv(path)

    required_cols = {"rule_id", "severity", "category", "title"}
    missing = required_cols - set(df.columns)
    if missing:
        console.print(f"[red]Error:[/red] CSV is missing required columns: {missing}")
        console.print("[dim]Hint: Use stig_parser.py --export to generate this file[/dim]")
        sys.exit(1)

    return df


def enrich_findings(df: pd.DataFrame, poc: str, scan_date: datetime) -> pd.DataFrame:
    """Add all POA&M fields to the findings DataFrame."""

    df = df.copy()

    # Sort by severity: CAT I first, then II, then III
    df = df.sort_values(
        by="severity",
        key=lambda x: x.map({"high": 0, "medium": 1, "low": 2}),
        ignore_index=True
    )

    # Sequential weakness IDs
    df["weakness_id"] = [f"POAM-{str(i + 1).zfill(3)}" for i in range(len(df))]

    # Calculate scheduled completion date from scan date + remediation window
    df["scheduled_completion"] = df["severity"].apply(
        lambda s: (
            scan_date + timedelta(days=REMEDIATION_DAYS.get(s, 90))
        ).strftime("%Y-%m-%d")
    )

    df["point_of_contact"]  = poc
    df["resources_required"] = "TBD"
    df["milestone"]          = "Remediation in progress"
    df["milestone_changes"]  = "None"
    df["source"]             = "STIG"
    df["status"]             = "Open"

    return df


# ── Workbook layer ────────────────────────────────────────────────────────────

def style_header_row(ws, row: int, num_cols: int):
    """Apply dark navy header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER


def create_cover_sheet(wb: Workbook, system_name: str, df: pd.DataFrame, scan_date: datetime):
    """Build a summary cover sheet with system metadata and finding counts."""

    ws = wb.active
    ws.title = "Cover Sheet"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40

    # Title banner
    ws.merge_cells("A1:B1")
    title_cell            = ws["A1"]
    title_cell.value      = "PLAN OF ACTION AND MILESTONES (POA&M)"
    title_cell.fill       = COVER_FILL
    title_cell.font       = WHITE_FONT
    title_cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Metadata rows
    metadata = [
        ("System Name",     system_name),
        ("Scan Date",       scan_date.strftime("%Y-%m-%d")),
        ("Date Generated",  datetime.today().strftime("%Y-%m-%d")),
        ("Source",          "DISA STIG"),
        ("Total Findings",  str(len(df))),
        ("CAT I  (High)",   str(len(df[df["severity"] == "high"]))),
        ("CAT II (Medium)", str(len(df[df["severity"] == "medium"]))),
        ("CAT III (Low)",   str(len(df[df["severity"] == "low"]))),
        ("Open Items",      str(len(df[df["status"] == "Open"]))),
    ]

    for row_idx, (label, value) in enumerate(metadata, start=2):
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        value_cell = ws.cell(row=row_idx, column=2, value=value)

        label_cell.font      = BOLD_FONT
        value_cell.font      = NORMAL_FONT
        label_cell.border    = THIN_BORDER
        value_cell.border    = THIN_BORDER
        label_cell.alignment = Alignment(vertical="center")
        value_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row_idx].height = 20

        # Color code the severity count rows
        if "(High)" in label:
            value_cell.fill = FILL_COLORS["high"]
        elif "(Medium)" in label:
            value_cell.fill = FILL_COLORS["medium"]
        elif "(Low)" in label:
            value_cell.fill = FILL_COLORS["low"]


def create_findings_sheet(wb: Workbook, df: pd.DataFrame):
    """Build the main POA&M findings worksheet."""

    ws = wb.create_sheet(title="POA&M Findings")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"  # Keep header visible while scrolling

    # Write headers
    headers = [col[0] for col in POAM_COLUMNS]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, row=1, num_cols=len(POAM_COLUMNS))

    # Set column widths
    for col_idx, (_, width) in enumerate(POAM_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        severity = row.severity.lower()



        row_data = [
            row.weakness_id,
            row.title,
            row.point_of_contact,
            row.resources_required,
            row.scheduled_completion,
            row.milestone,
            row.milestone_changes,
            row.source,
            row.status,
            row.category,
            row.severity.upper(),
            row.rule_id,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply color only to Category (col 10) and Severity (col 11)
            if col_idx in (10, 11):
                cell.fill = FILL_COLORS.get(severity, PatternFill())
            else:
                cell.fill = PatternFill()

            cell.font      = NORMAL_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


        ws.row_dimensions[row_idx].height = 30


# ── CLI entry point ───────────────────────────────────────────────────────────

@click.command()
@click.argument("filepath", type=click.Path())
@click.option("--system",    "-s", default="Unknown System",   help="Name of the system being assessed")
@click.option("--output",    "-o", default="poam_output.xlsx", help="Output filename for the Excel workbook")
@click.option("--poc",       "-p", default="Unassigned",       help="Point of contact for remediation")
@click.option("--scan-date", "-d", default=None,               help="STIG scan date YYYY-MM-DD (defaults to today)")
def main(filepath, system, output, poc, scan_date):
    """Generate a formatted POA&M Excel workbook from a STIG parser CSV.

    \b
    Examples:
      python3 poam_generator.py reports/poam_input.csv
      python3 poam_generator.py reports/poam_input.csv --system "RHEL 9" --output reports/RHEL9_POAM.xlsx
      python3 poam_generator.py reports/poam_input.csv -s "RHEL 9" -p "John Smith" -d 2026-03-02
    """

    console.print(f"\n[bold blue]POA&M Generator[/bold blue] — Reading [cyan]{filepath}[/cyan]\n")

    # Parse and validate scan date
    if scan_date:
        try:
            parsed_date = datetime.strptime(scan_date, "%Y-%m-%d")
        except ValueError:
            console.print("[red]Error:[/red] --scan-date must be in YYYY-MM-DD format")
            sys.exit(1)
    else:
        parsed_date = datetime.today()

    # Pipeline: load → enrich → build → save
    df = load_findings(filepath)
    df = enrich_findings(df, poc, parsed_date)

    wb = Workbook()
    create_cover_sheet(wb, system, df, parsed_date)
    create_findings_sheet(wb, df)

    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Terminal summary
    cat1 = len(df[df["severity"] == "high"])
    cat2 = len(df[df["severity"] == "medium"])
    cat3 = len(df[df["severity"] == "low"])

    summary = Table(show_header=False, box=None, padding=(0, 2))
    summary.add_column(style="bold")
    summary.add_column()
    summary.add_row("System",           system)
    summary.add_row("Total Findings",   str(len(df)))
    summary.add_row("[red]CAT I[/red]",         str(cat1))
    summary.add_row("[yellow]CAT II[/yellow]",  str(cat2))
    summary.add_row("[green]CAT III[/green]",   str(cat3))
    summary.add_row("Output",           str(output_path))

    console.print(Panel(summary, title="[green]POA&M Generated[/green]", border_style="green"))


if __name__ == "__main__":
    main()
